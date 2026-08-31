"""Per-block Skill Studio gradebook: GET /api/studio/reports/all + /api/studio/reports/export.xlsx"""
import io
import os

import pytest
import requests
from dotenv import dotenv_values

frontend_env = dotenv_values("/app/frontend/.env")
base_url = os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")
if not base_url:
    raise RuntimeError("REACT_APP_BACKEND_URL missing")
BASE_URL = base_url.rstrip("/") + "/api"

GUIDE_TOK = "em_guide_tok"
EXP_TOK = "em_exp_tok"
SUB_COLS = ["Lessons %", "Skills %", "Block Task %", "Checkpoint %"]
TRACKS = ["docs", "sheets", "slides", "email"]


@pytest.fixture(scope="module")
def guide():
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {GUIDE_TOK}"})
    return s


@pytest.fixture(scope="module")
def explorer():
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {EXP_TOK}"})
    return s


@pytest.fixture(scope="module")
def report(guide):
    r = guide.get(f"{BASE_URL}/studio/reports/all", timeout=60)
    assert r.status_code == 200, r.text[:400]
    return r.json()


# --- reports/all structure ---
class TestReportsAll:
    def test_auth_required(self):
        r = requests.get(f"{BASE_URL}/studio/reports/all", timeout=30)
        assert r.status_code in (401, 403), r.status_code

    def test_explorer_forbidden(self, explorer):
        r = explorer.get(f"{BASE_URL}/studio/reports/all", timeout=30)
        assert r.status_code == 403

    def test_shape(self, report):
        assert set(report.keys()) == {"block_meta", "students"}
        # old aggregate keys must be gone
        assert "totals" not in report
        for t in TRACKS:
            blocks = report["block_meta"][t]
            assert len(blocks) == 3, f"{t} has {len(blocks)} blocks"
            for i, b in enumerate(blocks):
                assert b["index"] == i + 1
                assert isinstance(b["cp"], str) and b["cp"]
                assert isinstance(b["lessons_total"], int) and b["lessons_total"] > 0
                assert isinstance(b["skills_total"], int)
                assert isinstance(b["has_task"], bool)

    def test_students_and_em_exp_email_block(self, report):
        students = report["students"]
        assert isinstance(students, list) and len(students) > 0
        ids = [s["user_id"] for s in students]
        assert "em-exp" in ids, f"em-exp not in guide roster: {ids}"
        for s in students:
            assert "totals" not in s
            for t, tr in s["tracks"].items():
                assert t in TRACKS
                assert isinstance(tr["writing_flag"], bool)
                assert len(tr["blocks"]) == 3
                for b in tr["blocks"]:
                    for grp in ("lessons", "skills"):
                        g = b[grp]
                        assert g["avg"] is None or isinstance(g["avg"], (int, float))
                        assert 0 <= g["done"] <= g["total"]
                        if g["done"] == 0:
                            assert g["avg"] is None, f"{s['user_id']}/{t}/B{b['index']}/{grp} avg set with 0 done"
                        else:
                            assert 0 <= g["avg"] <= 100
                    assert b["task"]["score"] is None or 0 <= b["task"]["score"] <= 100
                    assert isinstance(b["task"]["has"], bool)
                    assert b["checkpoint"]["score"] is None or 0 <= b["checkpoint"]["score"] <= 100
                    assert isinstance(b["checkpoint"]["passed"], bool)

    def test_em_exp_has_email_progress(self, report):
        s = next(x for x in report["students"] if x["user_id"] == "em-exp")
        assert "email" in s["tracks"], f"tracks={list(s['tracks'].keys())}"
        b1 = s["tracks"]["email"]["blocks"][0]
        print("em-exp email B1:", b1)
        # Block Task% is a single value, not an aggregate over 3 tasks
        assert b1["task"]["score"] is None or isinstance(b1["task"]["score"], (int, float))


# --- xlsx export ---
class TestXlsxExport:
    def test_explorer_forbidden(self, explorer):
        r = explorer.get(f"{BASE_URL}/studio/reports/export.xlsx", timeout=60)
        assert r.status_code == 403

    def test_headers_and_structure(self, guide, report):
        r = guide.get(f"{BASE_URL}/studio/reports/export.xlsx", timeout=90)
        assert r.status_code == 200, r.text[:300]
        assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in r.headers.get("content-disposition", "")
        assert ".xlsx" in r.headers.get("content-disposition", "")
        assert r.content[:2] == b"PK"

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.freeze_panes == "C4"
        assert ws["A1"].value == "Explorer"
        assert ws["B1"].value == "Email"

        merged = {str(m) for m in ws.merged_cells.ranges}
        # row1: 4 track headers each merged across 12 cols
        row1 = [ws.cell(row=1, column=c).value for c in range(3, 3 + 48)]
        labels = [v for v in row1 if v]
        assert labels == ["Docs", "Sheets", "Slides", "Email"], labels
        # row2: Block 1/2/3 per track, merged across 4 columns
        row2 = [ws.cell(row=2, column=c).value for c in range(3, 3 + 48)]
        assert [v for v in row2 if v] == ["Block 1", "Block 2", "Block 3"] * 4
        for start in range(3, 51, 4):
            a = openpyxl.utils.get_column_letter(start)
            b = openpyxl.utils.get_column_letter(start + 3)
            assert f"{a}2:{b}2" in merged, f"missing block merge {a}2:{b}2"
        # row3: repeating sub-columns
        row3 = [ws.cell(row=3, column=c).value for c in range(3, 51)]
        assert row3 == SUB_COLS * 12, row3

        # data rows
        n = len(report["students"])
        assert ws.max_row == 3 + n, f"max_row={ws.max_row} students={n}"

        # CRITICAL: numeric or blank, never a string like "10/14"
        bad = []
        numeric_found = 0
        for row in ws.iter_rows(min_row=4, min_col=3, max_col=50):
            for c in row:
                if c.value is None:
                    continue
                if isinstance(c.value, str):
                    bad.append((c.coordinate, c.value))
                else:
                    numeric_found += 1
                    assert 0 <= c.value <= 100, (c.coordinate, c.value)
        assert not bad, f"string/date-corruptible cells: {bad[:10]}"
        assert numeric_found > 0, "no numeric grade cells at all — export looks empty"
        print(f"xlsx OK: {n} students, {numeric_found} numeric cells")

    def test_expedition_filter(self, guide):
        exps = guide.get(f"{BASE_URL}/expeditions", timeout=30)
        assert exps.status_code == 200
        eid = exps.json()[0]["expedition_id"]
        r = guide.get(f"{BASE_URL}/studio/reports/all", params={"expedition_id": eid}, timeout=60)
        assert r.status_code == 200
        assert len(r.json()["students"]) > 0
        x = guide.get(f"{BASE_URL}/studio/reports/export.xlsx", params={"expedition_id": eid}, timeout=90)
        assert x.status_code == 200 and x.content[:2] == b"PK"

    def test_bad_expedition_404(self, guide):
        r = guide.get(f"{BASE_URL}/studio/reports/all", params={"expedition_id": "nope-123"}, timeout=30)
        assert r.status_code == 404


def test_null_score_attempt_does_not_break_report(guide):
    """Regression: a completed assessment_attempt with score=null must NOT 500 the report
    (it previously threw 'NoneType > int' and the UI showed 'No activity')."""
    import asyncio
    from dotenv import load_dotenv
    load_dotenv()
    from motor.motor_asyncio import AsyncIOMotorClient

    async def _seed():
        c = AsyncIOMotorClient(os.environ["MONGO_URL"]); db = c[os.environ["DB_NAME"]]
        await db.assessment_attempts.insert_many([
            {"user_id": "em-exp", "assessment_id": "email-cp1", "status": "completed", "score": 90, "passed": True, "_seed_null": 1},
            {"user_id": "em-exp", "assessment_id": "email-cp1", "status": "completed", "score": None, "passed": False, "_seed_null": 1},
        ])

    async def _cleanup():
        c = AsyncIOMotorClient(os.environ["MONGO_URL"]); db = c[os.environ["DB_NAME"]]
        await db.assessment_attempts.delete_many({"_seed_null": 1})

    asyncio.get_event_loop().run_until_complete(_seed())
    try:
        r = guide.get(f"{BASE_URL}/studio/reports/all", timeout=60)
        assert r.status_code == 200, r.text[:400]
        assert len(r.json()["students"]) >= 1
        rx = guide.get(f"{BASE_URL}/studio/reports/export.xlsx", timeout=60)
        assert rx.status_code == 200
    finally:
        asyncio.get_event_loop().run_until_complete(_cleanup())
